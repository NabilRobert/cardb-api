"""
excel_to_db.py

Parses the DSS Motor inventory Excel workbook and loads it into the
`uploads` / `vehicles` tables (see schema.sql) in Postgres.

Handles the messy real-world quirks found in the source file:
  - "Posisi Unit" -> status (available / booked) + who booked it
  - "Keterangan" -> kept as raw notes, with a best-effort area guess
  - brand name typos (e.g. "Mitsubsihi" -> "Mitsubishi")
  - "Harga CASH" == 0 means "not offered for cash", not free -> stored as NULL
  - rows broken by Excel #REF! errors are skipped and reported, not inserted
  - Indonesian date strings ("20 Maret 2027") are parsed into real dates

Usage:
    python excel_to_db.py path/to/file.xlsx              # parse + insert into DB
    python excel_to_db.py path/to/file.xlsx --dry-run     # parse only, print results, no DB writes

Requires DATABASE_URL in a .env file alongside this script (or in the environment).
"""

import os
import re
import sys
import argparse
from datetime import date

import openpyxl
from dotenv import load_dotenv

load_dotenv()

# ---------------------------------------------------------------------------
# Reference data / normalization tables
# ---------------------------------------------------------------------------

# Known brand typos -> canonical name. Extend this as new variants show up.
BRAND_FIXES = {
    "mitsubsihi": "Mitsubishi",
    "mitsubishi": "Mitsubishi",
    "honda": "Honda",
    "toyota": "Toyota",
    "nissan": "Nissan",
    "daihatsu": "Daihatsu",
    "wuling": "Wuling",
    "hyundai": "Hyundai",
}

# Known area/branch codes that can appear at the tail of "Keterangan".
# Best-effort only -- if nothing matches, area is left NULL rather than guessed.
KNOWN_AREAS = [
    "TANGSEL", "JAKSEL", "JAKBAR", "JAKTIM", "JAKUT", "JAKPUS",
    "BOGOR", "BEKASI", "BKS", "TGR", "TANGERANG",
]

INDONESIAN_MONTHS = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10,
    "november": 11, "desember": 12,
}


def normalize_brand(raw):
    if not raw or not isinstance(raw, str):
        return None
    key = raw.strip().lower()
    return BRAND_FIXES.get(key, raw.strip().title())


def parse_indonesian_date(raw):
    """'20 Maret 2027' -> date(2027, 3, 20). Returns None if unparseable."""
    if not raw or not isinstance(raw, str):
        return None
    m = re.match(r"\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})\s*$", raw.strip())
    if not m:
        return None
    day, month_name, year = m.groups()
    month = INDONESIAN_MONTHS.get(month_name.strip().lower())
    if not month:
        return None
    try:
        return date(int(year), month, int(day))
    except ValueError:
        return None


def parse_status_and_reserved(posisi_unit):
    """
    'DSSM'          -> ('available', None)
    'BOOKED/CALVIN' -> ('booked', 'Calvin')
    '#REF!' or None -> (None, None)  -- caller should treat this row as broken
    """
    if not posisi_unit or not isinstance(posisi_unit, str):
        return None, None
    val = posisi_unit.strip()
    if val == "#REF!":
        return None, None
    if val.upper().startswith("BOOKED"):
        parts = val.split("/", 1)
        # Not title-cased on purpose: "PT" looks like a company/entity code,
        # not a person's name like "CALVIN" or "SEAN" -- preserve as written.
        reserved_by = parts[1].strip() if len(parts) > 1 else None
        return "booked", reserved_by
    if val.upper() == "DSSM":
        return "available", None
    # Unrecognized value -- keep it, but flag as-is rather than guessing
    return val.lower(), None


def parse_area(keterangan):
    """Best-effort: look for a known area code anywhere in the string."""
    if not keterangan or not isinstance(keterangan, str):
        return None
    upper = keterangan.upper()
    for area in KNOWN_AREAS:
        if area in upper:
            return area
    return None


def clean_plate(raw):
    if not raw or not isinstance(raw, str):
        return None
    val = raw.strip()
    return val if val and val != "#REF!" else None


def is_ref_error(*values):
    return any(v == "#REF!" for v in values)


# ---------------------------------------------------------------------------
# Sheet loaders -- each yields dicts ready to insert into `vehicles`
# ---------------------------------------------------------------------------

def load_pricelist_sheet(ws):
    """
    'Pricelist' sheet columns (row 6 is the header):
    No. | No. Pol | Merk | Tipe | Transmisi | Tahun | Warna | Km | Tgl STNK |
    Age | Harga CASH | Harga Kredit | Maksimal Disc Kredit | Kepemilikan |
    Keterangan | Posisi Unit | Total Nilai Stock
    """
    rows_out = []
    skipped = []

    for row_index, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8):
        (no, plate, merk, tipe, transmisi, tahun, warna, km, tgl_stnk, age,
         harga_cash, harga_kredit, max_disc, kepemilikan, keterangan,
         posisi_unit, total_nilai) = row[:17]

        if plate is None and merk is None:
            continue  # blank spacer row

        if is_ref_error(plate, merk, tipe, transmisi, tahun, warna, km, tgl_stnk):
            skipped.append({"sheet": "Pricelist", "row": row_index, "reason": "#REF! error"})
            continue

        status, reserved_by = parse_status_and_reserved(posisi_unit)

        rows_out.append({
            "license_plate": clean_plate(plate),
            "vin": None,
            "engine_no": None,
            "brand": normalize_brand(merk),
            "model_trim": tipe.strip() if isinstance(tipe, str) else tipe,
            "year": int(tahun) if isinstance(tahun, (int, float)) else None,
            "transmission": transmisi,
            "color": warna,
            "odometer_km": int(km) if isinstance(km, (int, float)) else None,
            "stnk_expiry_date": parse_indonesian_date(tgl_stnk),
            "status": status,
            "reserved_by": reserved_by,
            "location": "DSSM",  # this sheet is the DSS Motor Bintaro master list
            "ownership": kepemilikan,
            "price_cash": harga_cash if harga_cash not in (0, None) else None,
            "price_credit": harga_kredit if harga_kredit not in (0, None) else None,
            "max_credit_discount": None if max_disc == "#REF!" else max_disc,
            "notes_raw": keterangan if isinstance(keterangan, str) else None,
            "source": None,
            "sheet_name": "Pricelist",
            "row_index": row_index,
        })

    return rows_out, skipped


def load_smr_sheet(ws):
    """
    'SMR' sheet columns (row 3 is the header):
    No | Nomor Polisi | Merk | Tipe | Transmisi | Tahun | Warna | Km | Grade | Notes
    No status/price columns here -- treated as available stock at the SMR branch
    unless a future column says otherwise.
    """
    rows_out = []
    for row_index, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        no, plate, merk, tipe, transmisi, tahun, warna, km, grade, notes = row[:10]
        if plate is None and merk is None:
            continue

        rows_out.append({
            "license_plate": clean_plate(plate),
            "vin": None,
            "engine_no": None,
            "brand": normalize_brand(merk),
            "model_trim": tipe.strip() if isinstance(tipe, str) else tipe,
            "year": int(tahun) if isinstance(tahun, (int, float)) else None,
            "transmission": transmisi,
            "color": warna,
            "odometer_km": int(km) if isinstance(km, (int, float)) else None,
            "stnk_expiry_date": None,
            "status": "available",
            "reserved_by": None,
            "location": "SMR",
            "ownership": None,
            "price_cash": None,
            "price_credit": None,
            "max_credit_discount": None,
            "notes_raw": notes if isinstance(notes, str) else grade,
            "source": None,
            "sheet_name": "SMR",
            "row_index": row_index,
        })
    return rows_out, []


# ---------------------------------------------------------------------------
# DB insertion
# ---------------------------------------------------------------------------

def insert_into_db(filename, all_rows):
    import psycopg2

    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO uploads (filename) VALUES (%s) RETURNING id",
                    (filename,),
                )
                upload_id = cur.fetchone()[0]

                for r in all_rows:
                    cur.execute(
                        """
                        INSERT INTO vehicles (
                            license_plate, vin, engine_no, brand, model_trim, year,
                            transmission, color, odometer_km, stnk_expiry_date,
                            status, reserved_by, location, ownership,
                            price_cash, price_credit, max_credit_discount,
                            notes_raw, source, upload_id, sheet_name, row_index
                        ) VALUES (
                            %(license_plate)s, %(vin)s, %(engine_no)s, %(brand)s, %(model_trim)s, %(year)s,
                            %(transmission)s, %(color)s, %(odometer_km)s, %(stnk_expiry_date)s,
                            %(status)s, %(reserved_by)s, %(location)s, %(ownership)s,
                            %(price_cash)s, %(price_credit)s, %(max_credit_discount)s,
                            %(notes_raw)s, %(source)s, %(upload_id)s, %(sheet_name)s, %(row_index)s
                        )
                        """,
                        {**r, "upload_id": upload_id},
                    )
        print(f"Inserted {len(all_rows)} vehicles under upload_id={upload_id}")
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--dry-run", action="store_true", help="parse only, don't touch the database")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    all_rows = []
    all_skipped = []

    if "Pricelist" in wb.sheetnames:
        rows, skipped = load_pricelist_sheet(wb["Pricelist"])
        all_rows += rows
        all_skipped += skipped

    if "SMR" in wb.sheetnames:
        rows, skipped = load_smr_sheet(wb["SMR"])
        all_rows += rows
        all_skipped += skipped

    print(f"Parsed {len(all_rows)} vehicle rows, skipped {len(all_skipped)} broken rows.")
    for s in all_skipped:
        print("  skipped:", s)

    if args.dry_run:
        for r in all_rows[:10]:
            print(r)
        print("... (dry run, nothing written to the database)")
        return

    insert_into_db(os.path.basename(args.xlsx_path), all_rows)


if __name__ == "__main__":
    main()
